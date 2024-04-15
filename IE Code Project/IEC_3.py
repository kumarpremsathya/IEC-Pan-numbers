import json
import sys
import os
import cv2
import easyocr
import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from selenium.webdriver.common.action_chains import ActionChains

from selenium.common.exceptions import WebDriverException
import openpyxl
import mysql.connector
import re


# Function to handle captcha processing
def solve_captcha_icegate_part3(browser, iec_code):
    try:
        iec_input = browser.find_element(By.XPATH, '//*[@id="pan_no"]')    
        iec_input.send_keys(iec_code)
        time.sleep(3)
            
        captcha_text = process_captcha_icegate_part3(browser)
        print("captcha_text " , captcha_text)
        #  Input the recognized captcha text
        
        captcha_input = browser.find_element(By.XPATH, '//*[@id="captchaResp"]')
        captcha_input.clear()  # Clear the input field
        captcha_input.send_keys(captcha_text)

        time.sleep(3)

        # Click on the element with XPath '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dd/a'
        view_button = browser.find_element(By.XPATH, '//*[@id="SubB"]')
        view_button.click()
        time.sleep(3)

        # Check if the "Invalid Code! Please try again!" message is displayed
        try:
            error_message = browser.find_element(By.XPATH, '//*[@id="pagetable"]/tbody/tr[3]/td[2]/ul/li/span').text
            print("error_message :", error_message) 
            
        
            if "Invalid captcha! Please try again." in error_message:
                return solve_captcha_icegate_part3(browser, iec_code)  # Retry solving captcha recursively
                
            else:
                pass
        except Exception as e:
                print("Error in message :", e)
                return True
                
                 
                
    except Exception as e:
        print("Error :", e)
       



# Function to handle captcha processing
def process_captcha_icegate_part3(browser):
    # Get the captcha image element
    captcha_image = browser.find_element(By.XPATH, '//*[@id="capimg"]')

    # Save the captcha image locally
    with open('captcha.png', 'wb') as file:
        file.write(captcha_image.screenshot_as_png)

    # Load the reader
    reader = easyocr.Reader(['en'])
    result = reader.readtext('captcha.png', detail=0)
    captcha_text = ''.join(result)
    

    return captcha_text

# Disable printing progress information during download
def no_progress(blocknum, bs, size):
    pass

# # Redirect stdout to a file
# sys.stdout = open(os.devnull, 'w')





def scrape_data_icegate_part3(browser):
    global scraped_data_df  # Access the global DataFrame
    try:
        
    
        # Locate the element containing the data using its XPath
        data_element = browser.find_element(By.XPATH, '//*[@id="sub_content"]/div[2]')
        
        
        # Get the HTML content of the element
        data_html = data_element.get_attribute('innerHTML')
        print("data_html :", data_html)
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(data_html, 'html.parser')
        
    
        # Find the table
        table = soup.find('table', id='pagetable')
        print("table", table)
        
        # Extract table headers
        headers = [header.text.strip() for header in table.find_all('th')]

       # Extract table rows
        rows = []
        for row in table.find_all('tr'):
            data = [td.text.strip() for td in row.find_all('td')]
            if data and len(data) == len(headers):  # Ensure data length matches headers length
                row_dict = {headers[i]: data[i] for i in range(len(headers))}
                rows.append(row_dict)
            elif data:
                print("Mismatch in data length:", data)
        # Print the dictionary containing the last row's data
        print("row_dict:", rows[-1])
        print("row_dict====" , row_dict)
        print("rows===", rows)

        # Create DataFrame
        df = pd.DataFrame(rows, columns=headers)
        print("df===", df)
        # Write DataFrame to Excel
        df.to_excel('table_data.xlsx', index=False)
        
           
        return row_dict
        
            # Write "Total Number Of Branches" and "Branch details" to Excel
            # df_table2 = pd.DataFrame({'Total Number Of Branches': [total_branches_value], 'Branch details': [branch_details_json]})
            
            # Concatenate df_table1 and df_table2
            # combined_df = pd.concat([df_table1, df_table2], axis=1)

            # # Write the combined DataFrame to Excel
            # combined_df.to_excel('branch_data.xlsx', index=False)
            # print("Data written to 'branch_data.xlsx'")

            # return True

        # else:
        #     print("Second table not found.")
        #     return data, None, None

    except Exception as e:
        print("Error occurred while scraping data:", e)
        return None, None, None
    



def icegate_part3(iec_code):
    global scraped_data_df # Add this line to access the global variable
    chrome_options = webdriver.ChromeOptions()
    
    
    # Add preferences to clear cache
    # chrome_prefs = {}
    # chrome_prefs["profile.default_content_settings"] = {"images": 2, "plugins": 2, "popups": 2, "geolocation": 2, 
    #                                                     "notifications": 2, "auto_select_certificate": 2, "fullscreen": 2, 
    #                                                     "mouselock": 2, "mixed_script": 2, "media_stream": 2, 
    #                                                     "media_stream_mic": 2, "media_stream_camera": 2, "protocol_handlers": 2,
    #                                                     "ppapi_broker": 2, "automatic_downloads": 2, "midi_sysex": 2, 
    #                                                     "push_messaging": 2, "ssl_cert_decisions": 2, "metro_switch_to_desktop": 2, 
    #                                                     "protected_media_identifier": 2, "app_banner": 2, "site_engagement": 2, 
    #                                                     "durable_storage": 2}
    # chrome_prefs["profile.default_content_setting_values"] = {"cookies": 2, "images": 2, "javascript": 1, "plugins": 2, "popups": 2, 
    #                                                           "geolocation": 2, "notifications": 2, "auto_select_certificate": 2, 
    #                                                           "fullscreen": 2, "mouselock": 2, "mixed_script": 2, "media_stream": 2, 
    #                                                           "media_stream_mic": 2, "media_stream_camera": 2, "protocol_handlers": 2,
    #                                                           "ppapi_broker": 2, "automatic_downloads": 2, "midi_sysex": 2, 
    #                                                           "push_messaging": 2, "ssl_cert_decisions": 2, "metro_switch_to_desktop": 2, 
    #                                                           "protected_media_identifier": 2, "app_banner": 2, "site_engagement": 2, 
    #                                                           "durable_storage": 2}
    # chrome_options.add_experimental_option("prefs", chrome_prefs)
    
    browser = webdriver.Chrome(options=chrome_options)
    browser.maximize_window()  # Maximize the browser window
    browser.get('https://old.icegate.gov.in/EnqMod/USER_PANDetails_action')

    try:
        
   
            
        iec_input = browser.find_element(By.XPATH, '//*[@id="pan_no"]')    
        iec_input.send_keys(iec_code)
        time.sleep(3)
        
    
    
          # Attempt to solve the captcha
        try:
            captcha_solved_icegate_part3 = solve_captcha_icegate_part3(browser, iec_code)
            print(f"captcha_solved in part3 for IEC code: {iec_code}", captcha_solved_icegate_part3)    
        
        except Exception as e:
            traceback.print_exc()
            print("captcha_solved:", e)    
        
        time.sleep(5)

        if captcha_solved_icegate_part3:
            print("captcha_solved_icegate_part3======:", captcha_solved_icegate_part3)    
            # Wait for a few seconds before checking for the presence of the message
            time.sleep(3)
            
             # Check if the message "Details for this IEC Number is not available." exists
            try:
                global scraped_data_df  # Access the global DataFrame
                unavailable_message = browser.find_element(By.XPATH, '//*[@id="sub_content"]/div[2]/div/span').text
                print("unavailable_message:", unavailable_message)  
                if "No record found for this PAN number. User is not registered with ICEGATE." in unavailable_message:
                    
                    #  # Define the column names to be set as NULL
                    # null_columns = [
                    #     "IEC Number","IEC Issuance Date", "IEC Status", "DEL Status", "IEC Cancelled Date",
                    #     "IEC Suspended Date", "File Number", "File Date", "DGFT RA Office", "Category of Exporters",'Firm Name', 'Address',
                    #     "RCMC Details", "Nature of concern/Firm", 'Total Number Of Branches', 'Branch details', 'Address', 'PAN','Icegate ID','Role','Registered Since',
                    #     'Registered for Filing Services'
                    # ]
                    
                    # # Create a dictionary with NULL values for the specified columns
                    # null_data = {column: ["NULL"] for column in null_columns}
                    
                    #  # Create a new row in the DataFrame with the scraped data
                    null_columns = scraped_data_df.columns.tolist()  # Get all column names from the DataFrame
                    null_data = {column: 'NULL' for column in null_columns}
                    
                    null_row = {column: 'NULL' for column in scraped_data_df.columns}
                    new_row = {
                        **null_row,
                        'IEC Number' : iec_code,
                        'Part_Type': 'no part',
                        # 'IEC Issuance Date': 'NULL',
                        # 'IEC Status' : 'NULL',
                        # 'DEL Status': 'NULL',
                        # 'IEC Cancelled Date' : 'NULL',
                        # 'IEC Suspended Date': 'NULL',
                        # 'File Number': 'NULL',
                        # 'File Date': 'NULL',
                        # 'DGFT RA Office': 'NULL',
                        # 'Nature of concern/Firm': 'NULL',
                        # 'Category of Exporters' : 'NULL',
                        # 'Firm Name' : 'NULL',
                        # 'Address' : 'NULL',
                        # 'Branch details': 'NULL',
                        # 'RCMC Details' : 'NULL',
                        # 'PAN' : 'NULL',
                        # 'Total Number Of Branches': 'NULL',
        
                        # 'Icegate ID' : 'NULL',
                        # 'Role': 'NULL',
                        # 'Registered Since' : 'NULL',
                        # 'Registered for Filing Services': 'NULL',
                        
                    }
                    
                    scraped_data_df= pd.concat([scraped_data_df, pd.DataFrame([new_row])], ignore_index=True)
                    print("scraped_data_df_part3 no part ==", scraped_data_df)
                    
                       
                else:
                    pass
                  
                        
            except NoSuchElementException:
                    
                    # Proceed with scraping the data if the message doesn't exist
                    # Define the column names to be set as NULL
                    # null_columns = [
                    #     "IEC Number","IEC Issuance Date", "IEC Status", "DEL Status", "IEC Cancelled Date",
                    #     "IEC Suspended Date", "File Number", "File Date", "DGFT RA Office", "Category of Exporters",'Firm Name', 'Address',
                    #     "RCMC Details", "Nature of concern/Firm", 'Total Number Of Branches', 'Branch details', 'Address', 'PAN'
                    # ]
                    
                    # # Create a dictionary with NULL values for the specified columns
                    # null_data = {column: ["NULL"] for column in null_columns}
                    
                    
                    scraped_data = scrape_data_icegate_part3(browser)
                    if scraped_data is not None:
                        row_dict = scraped_data
                        # Create a new row in the DataFrame with the scraped data
                        
                        #  # Create a new row in the DataFrame with the scraped data
                        null_columns = scraped_data_df.columns.tolist()  # Get all column names from the DataFrame
                        null_data = {column: 'NULL' for column in null_columns}
                        
                        null_row = {column: 'NULL' for column in scraped_data_df.columns}
                        new_row = {
                            **null_row,
                            'IEC Number' : iec_code,
                            'Part_Type': 'part3',
                            # 'IEC Issuance Date': 'NULL',
                            # 'IEC Status' : 'NULL',
                            # 'DEL Status': 'NULL',
                            # 'IEC Cancelled Date' : 'NULL',
                            # 'IEC Suspended Date': 'NULL',
                            # 'File Number': 'NULL',
                            # 'File Date': 'NULL',
                            # 'DGFT RA Office': 'NULL',
                            # 'Nature of concern/Firm': 'NULL',
                            # 'Category of Exporters' : 'NULL',
                            # 'Firm Name' : 'NULL',
                            # 'Address' : 'NULL',
                            # 'Branch details': 'NULL',
                            # 'RCMC Details' : 'NULL',
                            # 'PAN' : 'NULL',
                            # 'Total Number Of Branches': 'NULL',
            
                            'Icegate ID' : row_dict.get('Icegate ID', ''),
                            'Role': row_dict.get('Role', ''),
                            'Registered Since' : row_dict.get('Registered Since', ''),
                            'Registered for Filing Services': row_dict.get('Registered for Filing Services', ''),
                            
                        }
                        
                        scraped_data_df= pd.concat([scraped_data_df, pd.DataFrame([new_row])], ignore_index=True)
                        print("scraped_data_df_part3_table ==", scraped_data_df)
                        # Write the scraped data to an Excel file

                                           
    except Exception as e:
        traceback.print_exc()
        print("wrong xpath", e)

    finally:
        browser.quit()

        # Restore stdout
        # sys.stdout = sys.__stdout__



#########################################################################




# Function to handle captcha processing
def solve_captcha_icegate_part2(browser, iec_code):
    try:
        iec_input = browser.find_element(By.XPATH, '//*[@id="searchIECode"]')    
        iec_input.send_keys(iec_code)
        time.sleep(3)
            
        captcha_text = process_captcha_icegate_part2(browser)
        print("captcha_text " , captcha_text)
        #  Input the recognized captcha text
        
        captcha_input = browser.find_element(By.XPATH, '//*[@id="captchaResp"]')
        captcha_input.clear()  # Clear the input field
        captcha_input.send_keys(captcha_text)

        time.sleep(3)

        # Click on the element with XPath '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dd/a'
        view_button = browser.find_element(By.XPATH, '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dd/a')
        view_button.click()
        time.sleep(3)

        # Check if the "Invalid Code! Please try again!" message is displayed
        try:
            error_message = browser.find_element(By.XPATH, '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dt[5]/ul/li/span').text
            print("error_message :", error_message) 
            
        
            if "Invalid Code! Please try again!" in error_message:
                return solve_captcha_icegate_part2(browser, iec_code)  # Retry solving captcha recursively
                
            else:
                pass
        except Exception as e:
                print("Error in message in recursive call part2:", e)
                return True
                
                 
                
    except Exception as e:
        print("Error in part 2 entering inputs and captcha :", e)
       



# Function to handle captcha processing
def process_captcha_icegate_part2(browser):
    # Get the captcha image element
    captcha_image = browser.find_element(By.XPATH, '//*[@id="capimg"]')

    # Save the captcha image locally
    with open('captcha.png', 'wb') as file:
        file.write(captcha_image.screenshot_as_png)

    # Load the reader
    reader = easyocr.Reader(['en'])
    result = reader.readtext('captcha.png', detail=0)
    captcha_text = ''.join(result)
    

    return captcha_text

# Disable printing progress information during download
def no_progress(blocknum, bs, size):
    pass

# # Redirect stdout to a file
# sys.stdout = open(os.devnull, 'w')





def scrape_data_icegate_part2(browser):
    global scraped_data_df  # Access the global DataFrame
    try:
        
    
        # Locate the element containing the data using its XPath
        data_element = browser.find_element(By.XPATH, '//*[@id="sub_content"]/div[2]')
        
        
        # Get the HTML content of the element
        data_html = data_element.get_attribute('innerHTML')
        print("data_html :", data_html)
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(data_html, 'html.parser')
        
                # Find the table
        table = soup.find('table', id='pagetable')
        print("table", table)
        
        
        # Extract the data from the table
        data = {}
        rows = table.find_all('tr')
        address_parts = []
        for row in rows[1:]:  # Skip the header row
            cols = row.find_all('td')
            cols = [col.text.strip() for col in cols]
            if cols:
                field = cols[0].replace('IE Code', 'IEC Number').replace('Name', 'Firm Name')
                value = ' '.join(cols[1:])
                if field in ['Address', '']:  # Combine address parts
                    address_parts.append(value)
                else:
                    data[field] = value

        # Join address parts into a single string
        data['Address'] = ', '.join(address_parts)
        # print("data :", data)

        # Create a DataFrame from the extracted data
        df_table1 = pd.DataFrame([data])
        
        
    
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(data_html, 'html.parser')

       
        
        # Find the second table
        tables = soup.find_all('table', id='pagetable')
        if len(tables) >= 2:
            table = tables[1]
            # print("Second table:", table)

            # Extract "Total Number Of Branches" value
            total_branches_row = table.find('th', colspan=True)
            if total_branches_row:
                total_branches_value = total_branches_row.text.split(':')[1].strip()
                print("Total Number Of Branches:", total_branches_value)

            # Extract data from the second table
            branch_details = []
            rows = table.find_all('tr')
            current_branch_detail = {}
            for row in rows:
                cols = row.find_all('td')
                cols = [col.text.strip() for col in cols]
                if cols:
                    if cols[0] == 'Branch Serial Number':
                        if current_branch_detail:
                            branch_details.append(current_branch_detail)
                        current_branch_detail = {'Branch Serial Number': cols[1]}
                    elif cols[0] == 'ADDRESS':
                        current_branch_detail['Address'] = ', '.join(cols[1:])

            # Append the last branch detail
            if current_branch_detail:
                branch_details.append(current_branch_detail)

            print("Branch details in part2:", branch_details)

            # Convert "Branch details" to JSON
            branch_details_json = json.dumps(branch_details)
            
            
            
            # Return the scraped data
            return data, total_branches_value, branch_details_json
        
            # Write "Total Number Of Branches" and "Branch details" to Excel
            # df_table2 = pd.DataFrame({'Total Number Of Branches': [total_branches_value], 'Branch details': [branch_details_json]})
            
            # Concatenate df_table1 and df_table2
            # combined_df = pd.concat([df_table1, df_table2], axis=1)

            # # Write the combined DataFrame to Excel
            # combined_df.to_excel('branch_data.xlsx', index=False)
            # print("Data written to 'branch_data.xlsx'")

            # return True

        else:
            print("Second table not found.")
            return data, None, None

    except Exception as e:
        print("Error occurred while scraping data in part2:", e)
        return None, None, None
    



# Replace 'path/to/your/file.xlsx' with the actual path to your Excel file
df = pd.read_excel('IEC_details.xlsx', dtype={'IEC_CODE': str})




def icegate_part2(iec_code):
    global scraped_data_df # Add this line to access the global variable
    chrome_options = webdriver.ChromeOptions()
    
    max_retries = 3  # Number of retries
    retry_delay = 5  # Delay in seconds between retries

    for retry in range(max_retries):
        try:
            browser = webdriver.Chrome(options=chrome_options)
            browser.maximize_window()  # Maximize the browser window
            browser.get('https://old.icegate.gov.in/EnqMod/')

            try:
                
                iec_input = browser.find_element(By.XPATH, '//*[@id="searchIECode"]')    
                iec_input.send_keys(iec_code)
                time.sleep(3)
                
            
            
                # Attempt to solve the captcha
                try:
                    captcha_solved_icegate_part2 = solve_captcha_icegate_part2(browser, iec_code)
                    print(f"captcha_solved in part2 for IEC code: {iec_code}", captcha_solved_icegate_part2)    
                
                except Exception as e:
                    traceback.print_exc()
                    print("captcha_solved:", e)    
                
                time.sleep(5)

                if captcha_solved_icegate_part2:
                # Wait for a few seconds before checking for the presence of the message
                    time.sleep(3)
                    
                    # Check if the message "Details for this IEC Number is not available." exists
                    try:
                        global scraped_data_df  # Access the global DataFrame
                        unavailable_message = browser.find_element(By.XPATH, '//*[@id="sub_content"]/div[2]').text
                        if "No Record Found" in unavailable_message:
                            icegate_part3(iec_code)
                        else:
                            # Proceed with scraping the data if the message doesn't exist
                            # Define the column names to be set as NULL
                            # null_columns = [
                            #     "IEC Issuance Date", "IEC Status", "DEL Status", "IEC Cancelled Date",
                            #     "IEC Suspended Date", "File Number", "File Date", "DGFT RA Office", "Category of Exporters",
                            #     "RCMC DETAILS", "Nature of concern/Firm"
                            # ]
                            
                            # # Create a dictionary with NULL values for the specified columns
                            # null_data = {column: ["NULL"] for column in null_columns}
                            null_columns = scraped_data_df.columns.tolist()  # Get all column names from the DataFrame
                            null_data = {column: 'NULL' for column in null_columns}
                            
                            null_row = {column: 'NULL' for column in scraped_data_df.columns}
                            
                            scraped_data = scrape_data_icegate_part2(browser)
                            if scraped_data is not None:
                                data, total_branches_value, branch_details_json = scraped_data
                                # Create a new row in the DataFrame with the scraped data
                                new_row = {
                                    **null_row,
                                    'IEC Number': str(data.get('IEC Number', '')),  # Convert IEC Number to string
                                    'Part_Type': 'part2',
                                    # 'IEC Issuance Date': 'NULL',
                                    'IEC Status' : data.get('IEC Status', ''),
                                    # 'DEL Status': 'NULL',
                                    # 'IEC Cancelled Date' : 'NULL',
                                    # 'IEC Suspended Date': 'NULL',
                                    # 'File Number': 'NULL',
                                    # 'File Date': 'NULL',
                                    # 'DGFT RA Office': 'NULL',
                                    # 'Nature of concern/Firm': 'NULL',
                                    # 'Category of Exporters': 'NULL',
                                    'Firm Name': data.get('Firm Name', ''),
                                    'Address': data.get('Address', ''),
                                    'Branch details': branch_details_json,
                                    # 'RCMC Details': 'NULL',
                                    'PAN' : data.get('PAN', ''),
                                    
                                    'Total Number Of Branches': total_branches_value,
                                    
                                    # 'Icegate ID':'NULL',
                                    # 'Role':'NULL',
                                    # 'Registered Since':'NULL',
                                    # 'Registered for Filing Services':'NULL',
                                }
                                
                                
                            

                                
                                # Create a new DataFrame to store scraped data
                                # scraped_data_df = pd.DataFrame(columns=['IEC Number' ,'IEC Issuance Date','IEC Status', 'DEL Status', 'IEC Cancelled Date', 'IEC Suspended Date', 'File Number','File Date','DGFT RA Office',
                                #                 'Nature of concern/Firm', 'Category of Exporters' , 'Address', 'Branch details', 'RCMC Details' ,'PAN', 'Total Number Of Branches'])
                                
                                # Create a DataFrame from the row data
                                # icegate_data_df = pd.DataFrame([new_row])
                                
                                # final_df = pd.concat([final_df, icegate_data_df], ignore_index=True)

                                # Return the scraped data DataFrame
                                # return icegate_data_df
                                scraped_data_df= pd.concat([scraped_data_df, pd.DataFrame([new_row])], ignore_index=True)
                                print("scraped_data_df_part2====", scraped_data_df)
                                
                                # Write the scraped data to an Excel file
                                # scraped_data_df.to_excel('scraped_data.xlsx', index=False)
                        
                                
                    except NoSuchElementException:
                        # If the element is not found, proceed with scraping the data
                        # scrape_data_icegate1(browser)
                        pass
                    
            except Exception as e:
                traceback.print_exc()
                print("wrong xpath", e)
               
            break    
        except WebDriverException as e:
            if "net::ERR_CONNECTION_CLOSED" in str(e):
                if retry < max_retries - 1:
                    print(f"Connection closed, retrying in {retry_delay} seconds... (Attempt {retry + 1}/{max_retries})")
                    time.sleep(retry_delay)
                else:
                    print(f"Maximum retries ({max_retries}) reached, skipping IEC code: {iec_code}")
                    return
            else:
                raise e
            
        finally:
            browser.quit()

        # # Restore stdout
        # sys.stdout = sys.__stdout__


####################################################################################################################



# Function to handle captcha processing
def solve_captcha_dgft_part1(browser):
    try:
        captcha_text = process_captcha_dgft_part1(browser)
        print("captcha_text " , captcha_text)
        
        captcha_input = browser.find_element(By.XPATH, '//*[@id="txt_Captcha"]')
        captcha_input.clear()
        captcha_input.send_keys(captcha_text)

        time.sleep(3)

        view_button = browser.find_element(By.XPATH, '//*[@id="viewIEC1"]')
        view_button.click()
        time.sleep(3)

        # Check if the "Please enter valid captcha code" message is displayed
        error_message = browser.find_element(By.XPATH, '//*[@id="incCaptcha"]').text
        print(" error_message",error_message) 
        
        if "Please enter valid captcha code" in error_message:
            return solve_captcha_dgft_part1(browser)  # Retry solving captcha recursively
            
        else:
            return True
        
    except Exception as e:
                traceback.print_exc()
                print("Error occurred captcha recursively in part 1 :", e)

#Function to handle captcha processing
def process_captcha_dgft_part1(browser):
    try:
        # Get the captcha image element
        captcha_image = browser.find_element(By.XPATH, '//*[@id="captcha"]')


        # Save the captcha image locally
        with open('captcha.png', 'wb') as file:
            file.write(captcha_image.screenshot_as_png)
            
        # Preprocess the captcha image
        captcha_image = cv2.imread('captcha.png')
        gray_image = cv2.cvtColor(captcha_image, cv2.COLOR_BGR2GRAY)
        # _, binary_image = cv2.threshold(gray_image, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)
        # blurred_image = cv2.GaussianBlur(binary_image, (3, 3), 0)
        cv2.imwrite('preprocessed_captcha.png', gray_image)
        
        # Load the reader
        reader = easyocr.Reader(['en'])
        result = reader.readtext('captcha.png', detail=0)
        captcha_text = ''.join(result)
        print("captcha_text", captcha_text)

        return captcha_text
    
    except Exception as e:
                traceback.print_exc()
                print("Error occurred captcha process in part 1 :", e)
               
import csv


# Create a new DataFrame to store scraped data
scraped_data_df = pd.DataFrame(columns=['IEC Number', 'Part_Type', 'IEC Issuance Date','IEC Status', 'DEL Status', 'IEC Cancelled Date', 'IEC Suspended Date', 'File Number','File Date','DGFT RA Office',
                                        'Nature of concern/Firm', 'Category of Exporters' ,'Firm Name', 'Address', 'Branch details', 'RCMC Details' ,'PAN', 
                                        'Total Number Of Branches','Icegate ID','Role','Registered Since','Registered for Filing Services'])


def scrape_data_dgft_part1(browser):
    global scraped_data_df  # Access the global DataFrame
    try:
        
       
        symbol_click =  browser.find_element(By.XPATH, '//*[@id="custom-accordion"]/div[2]/div[1]/a')
        
        symbol_click.click
        
         # Locate the element containing the data using its XPath
        rcmc_element = browser.find_element(By.XPATH, '//*[@id="rcmc"]/div')
        # print("rcmc_element :" , rcmc_element)
        
         # Get the HTML content of the element
        rcmc_html = rcmc_element.get_attribute('innerHTML')
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(rcmc_html, 'html.parser')
        
        
        rcmc_branch_details = [] 
        
        # Scraping data from table format
        table = soup.find('table', class_='table table-hover custom-datatable')
        # print("table :" , table)
        if table:
            headers = [th.text.strip() for th in table.find_all('th')]
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if cells:
                    # Extract text from cells and add to all_branch_details
                    row_details = [cell.text.strip() for cell in cells]
                    rcmc_branch_details.append(dict(zip(headers, row_details)))
        
        # Convert all_branch_details to JSON string
        rcmc_details_json = json.dumps(rcmc_branch_details)

        # Create a Pandas DataFrame with the JSON data
        rcmc_df = pd.DataFrame({"RCMC DETAILS": [rcmc_details_json]})
        print("rcmc_df", rcmc_df)
        
        
        # Locate the element containing the data using its XPath
        data_element = browser.find_element(By.XPATH, '//*[@id="iecdetails"]')
        
        # Get the HTML content of the element
        data_html = data_element.get_attribute('innerHTML')
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(data_html, 'html.parser')

        # Scraping data from div elements with multiple rows
        div_elements = soup.find_all('div', class_='card-body')
                      
        div_details = []
        for div_element in div_elements:
            rows = div_element.find_all('div', class_='form-group')
            details = {}
            for row in rows:
                label_element = row.find('label', class_='font-12 font-weight-semi-bold')
                if label_element:
                    label = label_element.text.strip()
                    value_element = row.find('p', class_='font-12 text-gray')
                    if value_element:
                        value = value_element.text.strip()
                        details[label] = value
            div_details.append(details)
            print("div_details    :" , div_details)
            print("details    :" , details)

        all_branch_details = []  # Initialize a list to store all branch details
        
        # Loop through each pagination link until there is no "Next" button
        while True:
            # Locate the element containing the data using its XPath
            data_element = browser.find_element(By.XPATH, '//*[@id="iecdetails"]')
            
            # Get the HTML content of the element
            data_html = data_element.get_attribute('innerHTML')
            
            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(data_html, 'html.parser')

            # Scraping data from table format
            table = soup.find('table', class_='table table-hover custom-datatable dataTable no-footer')
            if table:
                headers = [th.text.strip() for th in table.find_all('th')]
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    if cells:
                        # Extract text from cells and add to all_branch_details
                        row_details = [cell.text.strip() for cell in cells]
                        # Loop through each cell in the row
                        for i in range(len(row_details)):
                            # Apply substitution to each string element
                            row_details[i] = re.sub(r'[\n\t]+', ' ', row_details[i])
                        print("row_details==", row_details)
                        all_branch_details.append(dict(zip(headers, row_details)))
            try:            
                 # Locate the "Next" button
                next_button = browser.find_element(By.ID, 'branchTable_next')
                # Scroll the "Next" button into view
                browser.execute_script("arguments[0].scrollIntoView();", next_button)
                # Click the "Next" button using JavaScript
                browser.execute_script("arguments[0].click();", next_button)
                
                # Wait for the table to load (you may need to implement this)
                # Add code to wait for the table to load here
                time.sleep(2)  # Adjust the delay according to your page loading time

                # Continue scraping data from the current page

                # Check if the "Next" button is disabled
                if 'disabled' in next_button.get_attribute('class'):
                    # If "Next" button is disabled, it means there are no more pages, so break out of the loop
                    break

                # Check if the ellipsis (...) is present in pagination
                ellipsis = browser.find_element(By.ID, 'branchTable_ellipsis')
                if 'disabled' not in ellipsis.get_attribute('class'):
                    # If ellipsis is present and not disabled, click on it to reveal more pages
                    ellipsis.click()
                    
                    # Wait for the table to load (you may need to implement this)
                    # Add code to wait for the table to load here
                    time.sleep(1)  # Adjust the delay according to your page loading time
            
                   # Continue with the rest of the scraping logic here
            except Exception as e:
                traceback.print_exc()
                print("Error occurred pagination branch details in part 1 :", e)
               
               
                  
        # Convert all_branch_details to JSON string
        branch_details_json = json.dumps(all_branch_details)
       
        print("branch_details_json=====", branch_details_json)
        
        # # Load JSON string into Python object
        # branch_details = json.loads(branch_details_json)
        #  # Convert dictionary to DataFrame
        # branch_details_df = pd.DataFrame(branch_details)
        
        # csv_file = 'branch_details.csv'
        # # Save DataFrame to CSV file
        # branch_details_df.to_csv(csv_file, index=False)
         
         # Return the scraped data
        # print("RCMC=========="  , rcmc_details_json, details, branch_details_json)
        
         # Define the column names to be set as NULL
        # null_columns = ["PAN", "Total Number Of Branches"]
        
        # # Create a dictionary with NULL values for the specified columns
        # null_data = {column: 'NULL' for column in null_columns}
        #  # Create a new row in the DataFrame with the scraped data
        null_columns = scraped_data_df.columns.tolist()  # Get all column names from the DataFrame
        null_data = {column: 'NULL' for column in null_columns}
         # Initialize a new row with NULL values for all columns
        null_row = {column: 'NULL' for column in scraped_data_df.columns}
        
        new_row = {
            **null_row ,
            'IEC Number': str(details.get('IEC Number', '')),  # Convert IEC Number to string
            'Part_Type':'part1',
            'IEC Issuance Date': details.get('IEC Issuance Date',''),
            'IEC Status' : details.get('IEC Status',''),
            'DEL Status': details.get('DEL Status',''),
            'IEC Cancelled Date' : details.get('IEC Cancelled Date',''),
            'IEC Suspended Date': details.get('IEC Suspended Date',''),
            'File Number': details.get('File Number',''),
            'File Date': details.get('File Date',''),
            'DGFT RA Office': details.get('DGFT RA Office',''),
            'Nature of concern/Firm': details.get('Nature of concern/Firm',''),
            'Category of Exporters' : details.get('Category of Exporters',''),
            'Firm Name' : details.get('Firm Name',''),
            'Address' : details.get('Address',''),
            'Branch details': branch_details_json,
            'RCMC Details' : rcmc_details_json,
            # # 'PAN' :'NULL', 
            # # 'Total Number Of Branches':'NULL',
            # 'Icegate ID':'NULL',
            # 'Role':'NULL',
            # 'Registered Since':'NULL',
            # 'Registered for Filing Services':'NULL',
            
            # 'PAN' : null_data.get('PAN',''),
            # 'Total Number Of Branches': null_data.get('Total Number Of Branches','')
            
        }
        
    
        scraped_data_df = pd.concat([scraped_data_df, pd.DataFrame([new_row])], ignore_index=True)
        print("scraped_data_df_part1====", scraped_data_df)
        return scraped_data_df 
        
        # return True
        # scraped_data_df.to_excel('scraped_data.xlsx', index=False)
        # return scraped_data_df 
        # print("scraped_data_df ", scraped_data_df)
        
        
        
        # return rcmc_details_json, details, branch_details_json
        # # Define the path to save the Excel file
        # excel_file = "scraped_data.xlsx"

        # # Create a Pandas DataFrame with the JSON data
        # branch_df = pd.DataFrame({"BRANCH DETAILS": [branch_details_json]})
        
        # # Write the DataFrame to an Excel file
        # branch_df.to_excel(excel_file, index=False)
        
        # print(f"Data has been exported to {excel_file}")
        
        # # Convert scraped data to DataFrames
        # table_df = pd.DataFrame(all_branch_details)
        # div_df = pd.DataFrame(div_details)
        
        # # Remove \n\t from DataFrame columns
        # table_df = table_df.replace(r'\n\t','', regex=True)
        # div_df = div_df.replace(r'\n\t','', regex=True)
        # branch_df = branch_df.replace(r'\n\t','', regex=True)
        # rcmc_df = rcmc_df.replace(r'\n\t','', regex=True)
        
        
        
        # # Define columns to set as NULL
        # null_columns = ["PAN", "Total Number Of Branches"]

        # # Set NULL values for specified columns
        # for column in null_columns:
        #     branch_df[column] = "NULL"
        
        

        # # Merge all DataFrames and write to Excel
        # merged_df = pd.concat([pd.DataFrame({"IEC NUMBER": ["0301014175"]}), div_df, branch_df , rcmc_df], axis=1)
        # # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        # #     merged_df.to_excel(writer, sheet_name='Merged Data', index=False)
            
            
        # # Check if the Excel file exists
        # if os.path.isfile(excel_file):
        #     # Load the existing Excel file
        #     book = load_workbook(excel_file)
        #     writer = pd.ExcelWriter(excel_file, engine='openpyxl', mode='a')

        #     # Check if the 'Merged Data' sheet exists
        #     if 'Merged Data' in book.sheetnames:
        #         # Overwrite the existing 'Merged Data' sheet
        #         with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        #             merged_df.to_excel(writer, sheet_name='Merged Data', index=False, startrow=0)
        #     else:
        #         # Create a new 'Merged Data' sheet
        #         with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        #             merged_df.to_excel(writer, sheet_name='Merged Data', index=False)
        # else:
        #     # Create the Excel file and write data to it
        #     merged_df.to_excel(excel_file, sheet_name='Merged Data', index=False)

        # print("merged_df :", merged_df)

        # print(f"Data has been exported to {excel_file}")
            
        # return True

    except Exception as e:
        traceback.print_exc()
        print("Error occurred while scraping data in part 1:", e)
        return False
    




    
# Disable printing progress information during download
def no_progress(blocknum, bs, size):
    pass




# Replace 'path/to/your/file.xlsx' with the actual path to your Excel file
df = pd.read_excel('IEC_details.xlsx', dtype={'IEC_CODE': str})










# Redirect stdout to a file
# sys.stdout = open(os.devnull, 'w')
def dgft_part1(iec_code, firm_name):
    global scraped_data_df  # Access the global DataFrame
    
    chrome_options = webdriver.ChromeOptions()
    browser = webdriver.Chrome(options=chrome_options)
    browser.maximize_window()  # Maximize the browser window
    browser.get('https://www.dgft.gov.in/CP/?opt=view-any-ice')

    try:
        next_xpath = browser.find_element(By.XPATH, '//*[@id="mainSectionWrap"]/div[3]/div/div[2]/div[1]')
        next_xpath.click()
        time.sleep(2)

        iec_input = browser.find_element(By.XPATH, '//*[@id="iecNo"]')
        entity_input = browser.find_element(By.XPATH, '//*[@id="entity"]')
        iec_input.send_keys(iec_code)
        time.sleep(3)
        entity_input.send_keys(firm_name)
        time.sleep(3)

        
        # Attempt to solve the captcha
        try:
            captcha_solved_dgft_part1 = solve_captcha_dgft_part1(browser)
            print(f"captcha_solved in part1 for IEC code: {iec_code}", captcha_solved_dgft_part1)    
        
        except Exception as e:
            traceback.print_exc()
            print("captcha_solved:", e)    
        
        time.sleep(5)
        
        if captcha_solved_dgft_part1:
            # Wait for a few seconds before checking for the presence of the message
            time.sleep(3)
            
            # Check if the message "Details for this IEC Number is not available." exists
            try:
                unavailable_message = browser.find_element(By.XPATH, '/html/body/div[16]/div/div/div/div[1]').text
                if "Details for this IEC Number is not available." in unavailable_message:
                    icegate_part2(iec_code)  # Call the icegate function if the message exists
                    # pass
                    
            
                else:
                     # Proceed with scraping the data if the message doesn't exist
                     # Define the column names to be set as NULL
                     pass
            except NoSuchElementException:
                traceback.print_exc()
                # If the element is not found, proceed with scraping the data
                scrape_data_dgft_part1(browser)

    except Exception as e:
        traceback.print_exc()
        print("An error occurred in part 1:", e)

    finally:
        browser.quit()
    


import csv

# Define the path to the Excel file
excel_file = 'scraped_data.xlsx'

# # Define the path to the CSV file
# csv_file = 'scraped_data.csv'

if os.path.isfile(excel_file):
    # Load the existing data into a DataFrame
    existing_data_df = pd.read_excel(excel_file, dtype={'IEC Number': str})
    print("existing_data_df1 :", existing_data_df)
    # Convert the 'IEC Number' column to string data type
    existing_data_df = existing_data_df.astype({'IEC Number': str})
    print("existing_data_df 2:", existing_data_df)
else:
    # Create an empty DataFrame if the file doesn't exist
    existing_data_df = pd.DataFrame(columns=scraped_data_df.columns)
    existing_data_df = existing_data_df.astype({'IEC Number': str})
    print("existing_data_df 3:", existing_data_df)
    
    



# def write_to_excel(data_df):
#     # Load the existing workbook or create a new one
#     workbook = openpyxl.load_workbook(excel_file) if os.path.isfile(excel_file) else openpyxl.Workbook()
#     worksheet = workbook.active

#     # Get the existing data from the worksheet
#     existing_data = []
#     for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
#         existing_data.append(row)

#     # Convert the existing data to a DataFrame (for checking if headers exist)
#     existing_headers_df = pd.DataFrame(existing_data, columns=data_df.columns)

#     # Write column headers if they don't already exist in the worksheet
#     if existing_headers_df.empty:
#         worksheet.append(data_df.columns.tolist())

#     # Get the existing data from the worksheet (after appending headers)
#     existing_data = []
#     for row in worksheet.iter_rows(min_row=2, values_only=True):
#         existing_data.append(row)

#     # Convert the existing data to a DataFrame
#     existing_df = pd.DataFrame(existing_data, columns=data_df.columns)

#     # Concatenate the existing data with the new data
#     combined_df = pd.concat([existing_df, data_df], ignore_index=True)

#     # Drop duplicates based on the "IEC Number" column
#     combined_df.drop_duplicates(subset="IEC Number", keep="last", inplace=True)

#     # Clear the existing worksheet data
#     worksheet.delete_rows(2, worksheet.max_row)

#     # Write the combined data to the worksheet
#     for row_num, row_data in enumerate(combined_df.values.tolist(), start=2):
#         for col_num, cell_value in enumerate(row_data, start=1):
#             worksheet.cell(row=row_num, column=col_num, value=str(cell_value))

#     # Save the workbook
#     workbook.save(excel_file)







import mysql.connector

# Connect to MySQL database
def connect_to_mysql():
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="root",
            database="mca"
        )
        print("Connected to MySQL database")
        return conn
    except mysql.connector.Error as e:
        print("Error connecting to MySQL database:", e)
        return None

# Function to insert scraped data into MySQL
def insert_into_mysql(conn, data_df):
    try:
        cursor = conn.cursor()

        # Define the SQL query to create the table
        create_table_query = """
        CREATE TABLE IF NOT EXISTS scraped_data (
            iec_number VARCHAR(255),
            part_type VARCHAR(255),
            iec_issuance_date DATE,
            iec_status VARCHAR(255),
            del_status VARCHAR(255),
            iec_cancelled_date DATE,
            iec_suspended_date DATE,
            file_number VARCHAR(255),
            file_date DATE,
            dgft_ra_office VARCHAR(255),
            nature_of_concern_firm VARCHAR(255),
            category_of_exporters VARCHAR(255),
            firm_name VARCHAR(255),
            address VARCHAR(255),
            branch_details TEXT,
            rcmc_details TEXT,
            pan VARCHAR(255),
            total_number_of_branches INT,
            icegate_id VARCHAR(255),
            role VARCHAR(255),
            registered_since DATE,
            registered_for_filing_services DATE,
            PRIMARY KEY (iec_number)
        )
        """

        # Execute the create table query
        cursor.execute(create_table_query)

        # Iterate over the rows in the DataFrame and execute the SQL query for each row
        for index, row in data_df.iterrows():
            
            # Check if the record already exists in the database
            cursor.execute("SELECT 1 FROM scraped_data WHERE iec_number = %s", (row['IEC Number'],))
            result = cursor.fetchone()

            # If the record does not exist, insert it into the database
            if not result:
            
            
                # Define the SQL query to insert data into the table
                sql = """
                INSERT INTO scraped_data(iec_number, part_type, iec_issuance_date, iec_status, del_status, iec_cancelled_date, iec_suspended_date, file_number,
                                              file_date, dgft_ra_office, nature_of_concern_firm, category_of_exporters, firm_name, address, branch_details, rcmc_details, pan,
                                              total_number_of_branches, icegate_id, role, registered_since, registered_for_filing_services)
                
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                
                # Extract row values and handle NULL values appropriately
                values = [
                    row['IEC Number'],
                    row['Part_Type'],
                    row['IEC Issuance Date'],
                    row['IEC Status'],
                    row['DEL Status'],
                    row['IEC Cancelled Date'], 
                    row['IEC Suspended Date'],
                    row['File Number'],
                    row['File Date'],
                    row['DGFT RA Office'],
                    row['Nature of concern/Firm'],
                    row['Category of Exporters'],
                    row['Firm Name'],
                    row['Address'],
                    row['Branch details'],
                    row['RCMC Details'],
                    row['PAN'],
                    row['Total Number Of Branches'],
                    row['Icegate ID'],
                    row['Role'],
                    row['Registered Since'] if not pd.isnull(row['Registered Since']) else None,
                    row['Registered for Filing Services'] if not pd.isnull(row['Registered for Filing Services']) else None,
                ]
                            
                cursor.execute(sql, values)
                print(f"Inserted record with IEC Number: {row['IEC Number']}")
            else:
                print(f"Skipped record with IEC Number {row['IEC Number']} as it already exists in the database")


        # Commit the transaction
        conn.commit()
        print("Data insertion into MySQL completed successfully")
    except mysql.connector.Error as e:
        print("Error inserting data into MySQL:", e)
        conn.rollback()


# Call this function after scraping and processing the data
def store_data_in_mysql(data_df):
    print("data for mysql====", data_df)
    conn = connect_to_mysql()
    if conn:
        insert_into_mysql(conn, data_df)
        conn.close()





def read_excel(start_index, end_index):
    global existing_data_df
    global scraped_data_df # Declare existing_data_df as global
    try:
        for row_index in range(start_index, end_index + 1):  # Add 1 to end_index to include it in the range
            # Read the specific row from the DataFrame
            row = df.iloc[row_index]

            # Extract IEC code from the row
            iec_code = str(row['IEC_CODE'])  # Convert IEC code to string
            print("iec_code:", iec_code)
            firm_name = row['FIRM NAME']
            print("firm_name:", firm_name)
            
            ## Call your function or code with the IEC code and firm name
            dgft_part1(iec_code, firm_name)
            
              
            
            # Concatenate the scraped data with the existing data
            combined_data_df = pd.concat([existing_data_df, scraped_data_df], ignore_index=True)
             
             # Fill NaN values with 'NULL' string
            combined_data_df = combined_data_df.fillna('NULL')
            
            combined_data_df.fillna('', inplace=True)
            # # Convert the 'IEC Number' column to string data type
            # combined_data_df = combined_data_df.astype({'IEC Number': str})
            
            
            
            # Replace '\n\t' with a single space
            combined_data_df.replace(r'\n\t','', regex=True)

            # Drop duplicates based on the "IEC Number" column
            combined_data_df.drop_duplicates(subset="IEC Number", keep="last", inplace=True)
            
            # Write the combined data to the Excel file using openpyxl
            # write_to_excel(combined_data_df)
            
            # combined_data_df.reset_index(drop=True, inplace=True)  # Reset index after dropping duplicates
            
            
            combined_data_df.to_excel(excel_file, index=False)
            print("combined_data_df===", combined_data_df)
            
            # Write the combined data to Excel
            # write_to_excel(combined_data_df)

            # Reset the existing_data_df with the combined data for the next iteration
            existing_data_df = combined_data_df.copy()

            # Reset the scraped_data_df DataFrame for the next iteration
            scraped_data_df = pd.DataFrame(columns=scraped_data_df.columns)
        
        # After processing all rows, store the combined data in MySQL
        store_data_in_mysql(combined_data_df)    
        
            
            
    except Exception as e:
        traceback.print_exc()
        print("Excel reading error:", e)

# Call read_excel with the desired range of row indices (3 to 5)
read_excel(4, 4)


